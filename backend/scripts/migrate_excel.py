"""一次性資料遷移工具：讀取舊版 index.html 匯出的 Excel 檔，寫入 PostgreSQL。

背景與決策記錄見 docs/requirements/2026-08-25-excel資料遷移.md。

用法（於 backend/ 目錄下執行）：
    python -m scripts.migrate_excel "路徑/肥尾日誌_2026-08-24.xlsx"

行為：
- 讀取「守宮」「餵食紀錄」兩個工作表（欄位對應舊版 index.html 匯出格式，見需求書）。
- 所有資料掛在 DEFAULT_USER_ID（目前系統唯一的固定 user）底下。
- Upsert，可安全重複執行：
    - 守宮用「名字」比對（舊版 uid() 不是合法 UUID，無法沿用舊 ID）。
    - 每日紀錄用 (gecko_id, date) 比對，對應 daily_logs 的 partial unique index。
- 照片：base64 → 解碼寫檔到 backend/uploads/geckos/<gecko_id>.<ext>，DB 只存相對路徑。
- 狀態：直接讀「狀態」欄位文字（進食/沒餵/拒食），不重新用 qty 推導：
    - 沒餵 → status=skipped，qty=null（非餵食日，qty 沒有意義）
    - 拒食 → status=refused，qty=0（有餵食但吃 0，qty=0 是有意義的資訊）
    - 進食 → status=fed，qty=Excel 原始數字
    - 不會產生 partial（舊資料沒有部分進食的判斷依據，符合既有需求書決策）
"""
from __future__ import annotations

import base64
import binascii
import re
import sys
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook

from app.constants import DEFAULT_USER_ID
from app.database import SessionLocal
from app.models import DailyLog, Gecko

BACKEND_DIR = Path(__file__).resolve().parent.parent
UPLOADS_DIR = BACKEND_DIR / "uploads" / "geckos"

GENDER_MAP = {"公": "male", "母": "female"}
SKIPPED_PATTERN = re.compile(r"沒餵|未餵|skip", re.IGNORECASE)
DATA_URI_PATTERN = re.compile(r"^data:image/(?P<ext>\w+);base64,(?P<data>.+)$", re.DOTALL)


def _cell(row: dict, *keys: str):
    for key in keys:
        if key in row and row[key] not in (None, ""):
            return row[key]
    return None


def _parse_date(value) -> date | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()
    for fmt in ("%Y-%m-%d", "%Y/%m/%d"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def _parse_number(value, default=0) -> float:
    if value is None or value == "":
        return default
    try:
        return float(value)
    except (TypeError, ValueError):
        return default


def _sheet_rows(ws) -> list[dict]:
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    header = [str(h).strip() if h is not None else "" for h in rows[0]]
    return [dict(zip(header, r)) for r in rows[1:] if any(v not in (None, "") for v in r)]


def _save_photo(gecko_id, data_uri: str | None) -> str | None:
    if not data_uri:
        return None
    match = DATA_URI_PATTERN.match(data_uri.strip())
    if not match:
        return None
    ext = match.group("ext").lower()
    if ext == "jpeg":
        ext = "jpg"
    try:
        raw = base64.b64decode(match.group("data"), validate=True)
    except (binascii.Error, ValueError):
        print(f"  [警告] 守宮 {gecko_id} 的照片 base64 解碼失敗，略過")
        return None

    UPLOADS_DIR.mkdir(parents=True, exist_ok=True)
    file_path = UPLOADS_DIR / f"{gecko_id}.{ext}"
    file_path.write_bytes(raw)
    return f"uploads/geckos/{gecko_id}.{ext}"


def _find_gecko(db, name: str) -> Gecko | None:
    return (
        db.query(Gecko)
        .filter(Gecko.user_id == DEFAULT_USER_ID, Gecko.name == name, Gecko.is_deleted.is_(False))
        .first()
    )


def _upsert_gecko_from_roster(db, name: str, morph, gender_text, birth, acquired, interval, note, photo_data_uri) -> Gecko:
    """處理「守宮」工作表的一列：這一列是該守宮的完整權威資料，全欄位覆寫。"""
    gecko = _find_gecko(db, name)
    if gecko is None:
        gecko = Gecko(user_id=DEFAULT_USER_ID, name=name)
        db.add(gecko)
        db.flush()  # 取得 gecko.id，photo 檔名需要用到

    gecko.morph = morph or None
    gecko.gender = GENDER_MAP.get(str(gender_text or "").strip(), "unknown")
    gecko.birth_date = _parse_date(birth)
    gecko.acquired_date = _parse_date(acquired)
    gecko.feeding_interval_days = max(1, min(60, int(_parse_number(interval, 7))))
    gecko.note = note or None

    photo_path = _save_photo(gecko.id, photo_data_uri)
    if photo_path:
        gecko.photo_path = photo_path

    return gecko


def _get_or_create_gecko_stub(db, name: str) -> Gecko:
    """處理「餵食紀錄」引用到、但「守宮」工作表沒有的名字：只用來讓 daily_logs 有
    地方掛，不能覆寫既有守宮的欄位（那些欄位的權威來源是「守宮」工作表那一列）。
    """
    gecko = _find_gecko(db, name)
    if gecko is None:
        gecko = Gecko(user_id=DEFAULT_USER_ID, name=name)
        db.add(gecko)
        db.flush()
    return gecko


def _derive_status_and_qty(status_text: str, qty_raw) -> tuple[str, int | None]:
    text = str(status_text or "")
    if SKIPPED_PATTERN.search(text):
        return "skipped", None
    qty = int(_parse_number(qty_raw, 0))
    if qty > 0:
        return "fed", qty
    return "refused", 0


def _upsert_daily_log(db, gecko: Gecko, log_date: date, food, size, status_text, qty_raw, poop_text, weight, note):
    status, qty = _derive_status_and_qty(status_text, qty_raw)
    row = (
        db.query(DailyLog)
        .filter(DailyLog.gecko_id == gecko.id, DailyLog.date == log_date, DailyLog.is_deleted.is_(False))
        .first()
    )
    is_new = row is None
    if row is None:
        row = DailyLog(gecko_id=gecko.id, date=log_date)
        db.add(row)

    row.status = status
    row.qty = qty
    row.food = food or None
    row.food_size = size or None
    row.poop = str(poop_text or "").strip() == "有"
    weight_val = _parse_number(weight, None) if weight not in (None, "") else None
    row.weight = weight_val
    row.note = note or None
    return is_new


def migrate(xlsx_path: Path) -> None:
    wb = load_workbook(xlsx_path, data_only=True)
    ws_gecko = wb["守宮"] if "守宮" in wb.sheetnames else wb[wb.sheetnames[0]]
    ws_logs = wb["餵食紀錄"] if "餵食紀錄" in wb.sheetnames else (wb[wb.sheetnames[1]] if len(wb.sheetnames) > 1 else None)

    gecko_rows = _sheet_rows(ws_gecko)
    log_rows = _sheet_rows(ws_logs) if ws_logs is not None else []

    db = SessionLocal()
    geckos_by_name: dict[str, Gecko] = {}
    new_geckos = updated_geckos = new_logs = updated_logs = skipped_rows = 0

    try:
        for row in gecko_rows:
            name = str(_cell(row, "名字", "名稱") or "").strip()
            if not name:
                skipped_rows += 1
                continue
            existed = name in geckos_by_name or _find_gecko(db, name) is not None
            gecko = _upsert_gecko_from_roster(
                db,
                name=name,
                morph=_cell(row, "品系"),
                gender_text=_cell(row, "性別"),
                birth=_cell(row, "出生日期"),
                acquired=_cell(row, "入手日期"),
                interval=_cell(row, "餵食頻率(天)"),
                note=_cell(row, "備註"),
                photo_data_uri=_cell(row, "照片(請勿編輯)", "照片"),
            )
            geckos_by_name[name] = gecko
            if existed:
                updated_geckos += 1
            else:
                new_geckos += 1

        for row in log_rows:
            log_date = _parse_date(_cell(row, "日期"))
            gecko_name = str(_cell(row, "守宮名字") or "").strip()
            if log_date is None or not gecko_name:
                skipped_rows += 1
                continue

            gecko = geckos_by_name.get(gecko_name)
            if gecko is None:
                pre_existing = _find_gecko(db, gecko_name) is not None
                gecko = _get_or_create_gecko_stub(db, gecko_name)
                geckos_by_name[gecko_name] = gecko
                if pre_existing:
                    updated_geckos += 1
                else:
                    new_geckos += 1

            is_new = _upsert_daily_log(
                db,
                gecko,
                log_date=log_date,
                food=_cell(row, "餌料"),
                size=_cell(row, "尺寸"),
                status_text=_cell(row, "狀態"),
                qty_raw=_cell(row, "進食數量"),
                poop_text=_cell(row, "排便"),
                weight=_cell(row, "體重(g)", "體重"),
                note=_cell(row, "備註"),
            )
            if is_new:
                new_logs += 1
            else:
                updated_logs += 1

        db.commit()
    finally:
        db.close()

    print(f"守宮：新增 {new_geckos}、更新 {updated_geckos}")
    print(f"每日紀錄：新增 {new_logs}、更新 {updated_logs}")
    if skipped_rows:
        print(f"略過（缺必要欄位）：{skipped_rows} 列")


if __name__ == "__main__":
    if len(sys.argv) != 2:
        print("用法：python -m scripts.migrate_excel <xlsx 檔案路徑>")
        sys.exit(1)
    path = Path(sys.argv[1])
    if not path.exists():
        print(f"找不到檔案：{path}")
        sys.exit(1)
    migrate(path)
